"""
Moteur de chargement et d'évaluation des questions depuis un fichier Excel.

Structure attendue de questions.xlsx :
  Feuille "Questions" avec colonnes :
    id        : identifiant unique (ex: Q001)
    question  : texte de la question
    type      : text | date | yes_no | choice | multiline | number
    options   : pour type=choice, valeurs séparées par "|"
    variable  : nom de la variable dans le template Word (ex: nom_defendeur)
    show_if   : condition d'affichage (ex: Q005=oui  ou  Q005!=non)  -- vide = toujours afficher
    section   : nom de la section/groupe (ex: "Parties", "Faits")
    required  : oui / non

  Feuille "Variants" (optionnelle) avec colonnes :
    condition     : ex: Q010=mineur
    template_file : nom du fichier .docx à utiliser (ex: template_mineur.docx)
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Dict, List, Optional

import openpyxl


# ---------------------------------------------------------------------------
# Modèle d'une question
# ---------------------------------------------------------------------------

@dataclass
class Question:
    id: str
    question: str
    type: str                          # text | date | yes_no | choice | multiline | number
    options: List[str] = field(default_factory=list)
    variable: str = ""
    show_if: str = ""
    section: str = "Général"
    required: bool = True

    # ------------------------------------------------------------------
    # Évaluation de la condition d'affichage
    # ------------------------------------------------------------------

    def should_show(self, answers: Dict[str, str]) -> bool:
        """Retourne True si la question doit être affichée selon les réponses actuelles."""
        expr = self.show_if.strip()
        if not expr:
            return True

        # Supporte plusieurs conditions reliées par AND (virgule ou espace+AND)
        # Ex : "Q005=oui, Q006!=non"
        sub_exprs = [e.strip() for e in re.split(r",|\bAND\b", expr, flags=re.IGNORECASE) if e.strip()]

        for sub in sub_exprs:
            if not self._eval_single(sub, answers):
                return False
        return True

    @staticmethod
    def _eval_single(expr: str, answers: Dict[str, str]) -> bool:
        """Évalue une condition atomique de la forme  QID=valeur  ou  QID!=valeur."""
        # Opérateur !=
        m = re.match(r"^(\w+)\s*!=\s*(.+)$", expr)
        if m:
            qid, val = m.group(1), m.group(2).strip().lower()
            return answers.get(qid, "").lower() != val

        # Opérateur =
        m = re.match(r"^(\w+)\s*=\s*(.+)$", expr)
        if m:
            qid, val = m.group(1), m.group(2).strip().lower()
            return answers.get(qid, "").lower() == val

        # Condition booléenne simple : "QID" → vrai si non vide
        return bool(answers.get(expr, "").strip())


# ---------------------------------------------------------------------------
# Moteur principal
# ---------------------------------------------------------------------------

class QuestionEngine:

    def __init__(self, excel_path: str) -> None:
        self.questions: List[Question] = []
        self.variants: List[Dict[str, str]] = []
        self._load(excel_path)

    # ------------------------------------------------------------------
    # Chargement depuis Excel
    # ------------------------------------------------------------------

    def _load(self, path: str) -> None:
        wb = openpyxl.load_workbook(path, data_only=True)

        # --- Feuille Questions ---
        if "Questions" not in wb.sheetnames:
            raise ValueError("Le fichier Excel doit contenir une feuille 'Questions'.")

        ws = wb["Questions"]
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            data = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}

            # Options pour type choice
            raw_opts = str(data.get("options", "") or "").strip()
            options = [o.strip() for o in raw_opts.split("|") if o.strip()] if raw_opts else []

            # required : oui par défaut
            req_raw = str(data.get("required", "oui") or "oui").strip().lower()
            required = req_raw not in ("non", "false", "0", "no")

            q = Question(
                id=str(data.get("id", "")).strip(),
                question=str(data.get("question", "") or "").strip(),
                type=str(data.get("type", "text") or "text").strip().lower(),
                options=options,
                variable=str(data.get("variable", "") or "").strip(),
                show_if=str(data.get("show_if", "") or "").strip(),
                section=str(data.get("section", "Général") or "Général").strip(),
                required=required,
            )
            if q.id and q.question:
                self.questions.append(q)

        # --- Feuille Variants (optionnelle) ---
        if "Variants" in wb.sheetnames:
            ws_v = wb["Variants"]
            v_headers = [str(c.value).strip() if c.value else "" for c in ws_v[1]]
            for row in ws_v.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                data = {v_headers[i]: str(row[i] or "").strip() for i in range(len(v_headers))}
                if data.get("condition") and data.get("template_file"):
                    self.variants.append(data)

    # ------------------------------------------------------------------
    # API publique
    # ------------------------------------------------------------------

    def get_visible_questions(self, answers: Dict[str, str]) -> List[Question]:
        """Retourne la liste des questions visibles selon les réponses actuelles."""
        return [q for q in self.questions if q.should_show(answers)]

    def get_template_for_answers(
        self, answers: Dict[str, str], default_template: str
    ) -> str:
        """
        Parcourt les règles de variants et retourne le fichier template
        correspondant à la première règle satisfaite.
        Retourne default_template si aucune règle ne correspond.
        """
        for variant in self.variants:
            cond = variant.get("condition", "")
            tmpl = variant.get("template_file", "")
            if not cond or not tmpl:
                continue
            # Réutilise la logique d'évaluation de Question
            dummy = Question(id="", question="", type="text", show_if=cond)
            if dummy.should_show(answers):
                return tmpl
        return default_template

    def get_sections(self) -> List[str]:
        """Retourne la liste ordonnée des sections (sans doublon)."""
        seen: List[str] = []
        for q in self.questions:
            if q.section not in seen:
                seen.append(q.section)
        return seen
